from __future__ import annotations

import math
from dataclasses import dataclass, field
from io import BytesIO
from threading import Lock
from typing import Any, Dict, Iterable, List, Optional, Tuple
from uuid import uuid4

import pandas as pd
from openpyxl.styles import PatternFill

TYPE_PRIORITY = ["integer", "float", "boolean", "date", "string"]
ALLOWED_TYPES = set(TYPE_PRIORITY)


@dataclass
class SessionData:
    rows: List[Dict[str, Any]]
    columns: List[str]
    detected_types: Dict[str, str]
    column_info: List[Dict[str, Any]]
    errors: List[Dict[str, Any]] = field(default_factory=list)
    duplicate_groups: List[List[int]] = field(default_factory=list)
    original_filename: str = "uploaded.xlsx"
    overrides: Dict[str, str] = field(default_factory=dict)
    workbook_bytes: bytes = b""
    sheet_name: str = ""
    sheet_names: List[str] = field(default_factory=list)


class SessionStore:
    def __init__(self) -> None:
        self._sessions: Dict[str, SessionData] = {}
        self._lock = Lock()

    def create_session(self, session: SessionData) -> str:
        session_id = str(uuid4())
        with self._lock:
            self._sessions[session_id] = session
        return session_id

    def get(self, session_id: str) -> SessionData:
        with self._lock:
            if session_id not in self._sessions:
                raise KeyError("Session not found")
            return self._sessions[session_id]

    def update(self, session_id: str, session: SessionData) -> None:
        with self._lock:
            if session_id not in self._sessions:
                raise KeyError("Session not found")
            self._sessions[session_id] = session


session_store = SessionStore()


def process_uploaded_file(file_bytes: bytes, filename: str) -> Tuple[SessionData, Dict[str, Any]]:
    excel_file = pd.ExcelFile(BytesIO(file_bytes))
    sheet_names = excel_file.sheet_names or ["Sheet1"]
    sheet_name = sheet_names[0]
    (
        rows,
        columns,
        detected_types,
        column_info,
        errors,
        duplicate_groups,
    ) = parse_sheet(file_bytes, sheet_name)
    session = SessionData(
        rows=rows,
        columns=columns,
        detected_types=detected_types,
        column_info=column_info,
        errors=errors,
        duplicate_groups=duplicate_groups,
        original_filename=filename or "uploaded.xlsx",
        workbook_bytes=file_bytes,
        sheet_name=sheet_name,
        sheet_names=sheet_names,
    )
    payload = {
        "columns": column_info,
        "rows": rows,
        "errors": errors,
        "duplicateGroups": duplicate_groups,
        "sheetName": sheet_name,
        "sheetNames": sheet_names,
    }
    return session, payload


def parse_sheet(
    file_bytes: bytes,
    sheet_name: Optional[str],
) -> Tuple[
    List[Dict[str, Any]],
    List[str],
    Dict[str, str],
    List[Dict[str, Any]],
    List[Dict[str, Any]],
    List[List[int]],
]:
    rows, columns = read_excel_into_rows(file_bytes, sheet_name=sheet_name)
    detected_types = detect_types(rows, columns)
    errors, duplicate_groups = validate_rows(rows, detected_types)
    column_info = [
        {
            "name": column,
            "detectedType": detected_types[column],
            "overrideType": None,
            "nullable": True,
        }
        for column in columns
    ]
    return rows, columns, detected_types, column_info, errors, duplicate_groups


def revalidate(
    session: SessionData,
    rows: List[Dict[str, Any]],
    overrides: Dict[str, str],
    column_info_payload: Optional[List[Dict[str, Any]]] = None,
) -> Dict[str, Any]:
    if column_info_payload:
        session.column_info = [
            {
                "name": column["name"],
                "detectedType": column.get("detectedType") or session.detected_types.get(column["name"], "string"),
                "overrideType": column.get("overrideType"),
                "nullable": column.get("nullable", True),
            }
            for column in column_info_payload
        ]
        session.columns = [column["name"] for column in session.column_info]
    coerced_overrides = {
        column: _normalize_type(value) for column, value in overrides.items() if column in session.columns
    }
    expected_types: Dict[str, str] = {}
    for column in session.column_info:
        column_name = column["name"]
        override_type = coerced_overrides.get(column_name) or column.get("overrideType")
        if override_type:
            column["overrideType"] = override_type
        detected_type = column.get("detectedType") or session.detected_types.get(column_name, "string")
        expected_types[column_name] = override_type or detected_type or "string"
    errors, duplicate_groups = validate_rows(rows, expected_types)
    session.rows = rows
    session.overrides = coerced_overrides
    session.errors = errors
    session.duplicate_groups = duplicate_groups
    session.detected_types = expected_types.copy()
    payload = {
        "columns": session.column_info,
        "rows": rows,
        "errors": errors,
        "duplicateGroups": duplicate_groups,
        "sheetName": session.sheet_name,
        "sheetNames": session.sheet_names,
    }
    return payload


def remove_rows(session: SessionData, row_ids: Iterable[int]) -> List[Dict[str, Any]]:
    removal_set = set(row_ids)
    filtered_rows = [row for row in session.rows if row["rowId"] not in removal_set]
    # Reassign rowId to keep ordering predictable
    for idx, row in enumerate(filtered_rows):
        row["rowId"] = idx
    session.rows = filtered_rows
    return filtered_rows


def generate_error_report(session: SessionData) -> bytes:
    return generate_error_report_from_rows(session.rows, session.errors)


def generate_error_report_from_rows(
    rows: List[Dict[str, Any]],
    errors: List[Dict[str, Any]],
) -> bytes:
    df = pd.DataFrame([row["values"] for row in rows])
    error_map = {(error["rowId"], error["column"]): error for error in errors}
    buffer = BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Data")
        workbook = writer.book
        worksheet = writer.sheets["Data"]
        red_fill = PatternFill(start_color="FFFFC7CE", end_color="FFFFC7CE", fill_type="solid")
        for (row_id, column), error in error_map.items():
            if column not in df.columns:
                continue
            col_idx = df.columns.get_loc(column) + 1
            cell = worksheet.cell(row=row_id + 2, column=col_idx)
            cell.fill = red_fill
        summary = pd.DataFrame(
            [
                {
                    "Row": error["rowId"] + 2,
                    "Column": error["column"],
                    "Expected Type": error["expectedType"],
                    "Value": error["actualValue"],
                    "Message": error["message"],
                }
                for error in errors
            ]
        )
        if not summary.empty:
            summary.to_excel(writer, index=False, sheet_name="Errors")
    buffer.seek(0)
    return buffer.getvalue()


def export_rows_to_excel(rows: List[Dict[str, Any]], columns: List[Dict[str, Any]]) -> bytes:
    ordered_columns = [column["name"] for column in columns] if columns else []
    data = []
    for row in rows:
        record = row["values"] if "values" in row else row
        if ordered_columns:
            record = {column: record.get(column) for column in ordered_columns}
        data.append(record)
    df = pd.DataFrame(data)
    if ordered_columns:
        missing = [column for column in ordered_columns if column not in df.columns]
        for column in missing:
            df[column] = None
        df = df[ordered_columns]
    buffer = BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Edited Data")
    buffer.seek(0)
    return buffer.getvalue()


def read_excel_into_rows(
    file_bytes: bytes,
    sheet_name: Optional[str] = None,
) -> Tuple[List[Dict[str, Any]], List[str]]:
    try:
        df = pd.read_excel(BytesIO(file_bytes), sheet_name=sheet_name)
    except ValueError as exc:
        raise ValueError("Unable to read the uploaded Excel file.") from exc
    original_columns = list(df.columns)
    columns = [str(column) for column in original_columns]
    na_mask = df.isna()
    rows: List[Dict[str, Any]] = []
    for idx in range(len(df)):
        row_series = df.iloc[idx]
        mask_series = na_mask.iloc[idx]
        normalized: Dict[str, Any] = {}
        for col_name, original in zip(columns, original_columns):
            if bool(mask_series[original]):
                normalized[col_name] = ""
            else:
                normalized[col_name] = _normalize_cell_value(row_series[original])
        rows.append({"rowId": idx, "values": normalized})
    return rows, columns


def detect_types(rows: List[Dict[str, Any]], columns: List[str]) -> Dict[str, str]:
    detected: Dict[str, str] = {}
    for column in columns:
        values = [row["values"].get(column) for row in rows]
        detected[column] = detect_column_type(values)
    return detected


def detect_column_type(values: Iterable[Any]) -> str:
    scores = {type_name: 0 for type_name in TYPE_PRIORITY}
    for value in values:
        if _is_null(value):
            continue
        if _looks_like_int(value):
            scores["integer"] += 1
            continue
        if _looks_like_float(value):
            scores["float"] += 1
            continue
        if _looks_like_bool(value):
            scores["boolean"] += 1
            continue
        if _looks_like_date(value):
            scores["date"] += 1
            continue
        scores["string"] += 1
    best_type = max(TYPE_PRIORITY, key=lambda type_name: (scores[type_name], -TYPE_PRIORITY.index(type_name)))
    if scores[best_type] == 0:
        return "string"
    return best_type


def validate_rows(rows: List[Dict[str, Any]], expected_types: Dict[str, str]) -> Tuple[List[Dict[str, Any]], List[List[int]]]:
    errors: List[Dict[str, Any]] = []
    for row in rows:
        for column, expected_type in expected_types.items():
            value = row["values"].get(column)
            if not _is_valid(value, expected_type):
                errors.append(
                    {
                        "rowId": row["rowId"],
                        "column": column,
                        "expectedType": expected_type,
                        "actualValue": value,
                        "message": f"Expected {expected_type}, received '{value}'",
                    }
                )
    duplicate_groups = identify_duplicates(rows, list(expected_types.keys()))
    return errors, duplicate_groups


def identify_duplicates(rows: List[Dict[str, Any]], columns: List[str]) -> List[List[int]]:
    seen: Dict[Tuple[Any, ...], List[int]] = {}
    for row in rows:
        key = tuple(_coerce_duplicate_value(row["values"].get(column)) for column in columns)
        seen.setdefault(key, []).append(row["rowId"])
    return [group for group in seen.values() if len(group) > 1]


def _normalize_cell_value(value: Any) -> Any:
    if hasattr(value, "item"):
        try:
            value = value.item()
        except Exception:
            pass
    if isinstance(value, pd.Timestamp):
        return value.isoformat()
    if isinstance(value, float) and math.isnan(value):
        return ""
    return value


def _is_null(value: Any) -> bool:
    if value is None:
        return True
    if isinstance(value, str):
        return value.strip() == ""
    if isinstance(value, float):
        return math.isnan(value)
    return False


def _looks_like_int(value: Any) -> bool:
    if isinstance(value, bool):
        return False
    if isinstance(value, int):
        return True
    if isinstance(value, float):
        return value.is_integer()
    if isinstance(value, str):
        value = value.strip()
        if value.startswith("-"):
            return value[1:].isdigit()
        return value.isdigit()
    return False


def _looks_like_float(value: Any) -> bool:
    if isinstance(value, bool):
        return False
    if isinstance(value, (int, float)):
        return True
    if isinstance(value, str):
        try:
            float(value)
            return True
        except ValueError:
            return False
    return False


def _looks_like_bool(value: Any) -> bool:
    if isinstance(value, bool):
        return True
    if isinstance(value, str):
        return value.strip().lower() in {"true", "false", "yes", "no", "0", "1"}
    if isinstance(value, (int, float)):
        return value in {0, 1}
    return False


def _looks_like_date(value: Any) -> bool:
    if isinstance(value, pd.Timestamp):
        return True
    if isinstance(value, str):
        if not value.strip():
            return False
        parsed = pd.to_datetime(value, errors="coerce")
        return parsed is not pd.NaT and not pd.isna(parsed)
    return False


def _is_valid(value: Any, expected_type: str) -> bool:
    if expected_type not in ALLOWED_TYPES:
        expected_type = "string"
    if _is_null(value):
        return True
    if expected_type == "string":
        return True
    if expected_type == "integer":
        return _looks_like_int(value)
    if expected_type == "float":
        return _looks_like_float(value)
    if expected_type == "boolean":
        return _looks_like_bool(value)
    if expected_type == "date":
        return _looks_like_date(value)
    return True


def _coerce_duplicate_value(value: Any) -> Any:
    if isinstance(value, str):
        return value.strip().lower()
    if isinstance(value, float) and math.isnan(value):
        return None
    return value


def _normalize_type(value: Optional[str]) -> Optional[str]:
    if value is None:
        return None
    normalized = value.lower()
    return normalized if normalized in ALLOWED_TYPES else None

